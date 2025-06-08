"""
Service d'export pour générer des rapports dans différents formats.
"""

from typing import List, Dict, Any, Optional
from datetime import datetime, timedelta
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from icalendar import Calendar, Event
import pytz
from sqlalchemy.orm import Session

from app.models.schedule import Schedule, ScheduleEntry
from app.models.teacher import Teacher
from app.models.subject import Subject
from app.models.class_group import ClassGroup
from app.models.room import Room


class ExportService:
    """Service pour exporter les emplois du temps dans différents formats."""
    
    def __init__(self, db: Session):
        self.db = db
    
    def export_schedule_to_excel(self, schedule_id: int) -> io.BytesIO:
        """Exporter un emploi du temps au format Excel."""
        schedule = self.db.query(Schedule).filter(Schedule.id == schedule_id).first()
        if not schedule:
            raise ValueError(f"Schedule {schedule_id} not found")
        
        # Créer un workbook
        wb = Workbook()
        
        # Créer une feuille par classe
        classes = self.db.query(ClassGroup).all()
        
        for idx, class_group in enumerate(classes):
            if idx == 0:
                ws = wb.active
                ws.title = class_group.code
            else:
                ws = wb.create_sheet(title=class_group.code)
            
            # Créer l'en-tête
            self._create_excel_header(ws, class_group)
            
            # Récupérer les entrées pour cette classe
            entries = self.db.query(ScheduleEntry).filter(
                ScheduleEntry.schedule_id == schedule_id,
                ScheduleEntry.class_id == class_group.id
            ).all()
            
            # Remplir l'emploi du temps
            self._fill_excel_schedule(ws, entries)
            
            # Appliquer le style
            self._apply_excel_style(ws)
        
        # Sauvegarder dans un buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def _create_excel_header(self, ws, class_group):
        """Créer l'en-tête de la feuille Excel."""
        # Titre
        ws.merge_cells('A1:G1')
        ws['A1'] = f"Emploi du temps - {class_group.name}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Jours de la semaine
        days = ['Heure', 'Dimanche', 'Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi']
        for col, day in enumerate(days, 1):
            cell = ws.cell(row=3, column=col, value=day)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
    
    def _fill_excel_schedule(self, ws, entries):
        """Remplir l'emploi du temps dans la feuille Excel."""
        # Heures de cours
        periods = [
            ('08:00-08:45', 4),
            ('08:50-09:35', 5),
            ('09:40-10:25', 6),
            ('10:40-11:25', 7),
            ('11:30-12:15', 8),
            ('12:20-13:05', 9),
            ('13:10-13:55', 10),
            ('14:00-14:45', 11)
        ]
        
        # Remplir les heures
        for time_slot, row in periods:
            ws.cell(row=row, column=1, value=time_slot).font = Font(bold=True)
        
        # Mapper les jours aux colonnes
        day_to_col = {
            'sunday': 2,
            'monday': 3,
            'tuesday': 4,
            'wednesday': 5,
            'thursday': 6,
            'friday': 7
        }
        
        # Remplir les cours
        for entry in entries:
            row = 4 + entry.period
            col = day_to_col.get(entry.day, 0)
            
            if col > 0:
                # Récupérer les informations
                subject = self.db.query(Subject).filter(Subject.id == entry.subject_id).first()
                teacher = self.db.query(Teacher).filter(Teacher.id == entry.teacher_id).first()
                room = self.db.query(Room).filter(Room.id == entry.room_id).first()
                
                # Créer le texte de la cellule
                cell_text = f"{subject.code}\n{teacher.code}\n{room.code}"
                cell = ws.cell(row=row, column=col, value=cell_text)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Colorer selon la matière
                if hasattr(subject, 'color') and subject.color:
                    cell.fill = PatternFill(start_color=subject.color[1:], end_color=subject.color[1:], fill_type="solid")
    
    def _apply_excel_style(self, ws):
        """Appliquer le style à la feuille Excel."""
        # Bordures
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Appliquer les bordures et ajuster les dimensions
        for row in ws.iter_rows(min_row=3, max_row=11, min_col=1, max_col=7):
            for cell in row:
                cell.border = thin_border
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 15
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 20
        
        # Ajuster la hauteur des lignes
        for row in range(4, 12):
            ws.row_dimensions[row].height = 40
    
    def export_schedule_to_pdf(self, schedule_id: int) -> io.BytesIO:
        """Exporter un emploi du temps au format PDF."""
        schedule = self.db.query(Schedule).filter(Schedule.id == schedule_id).first()
        if not schedule:
            raise ValueError(f"Schedule {schedule_id} not found")
        
        # Créer un buffer
        output = io.BytesIO()
        
        # Créer le document PDF
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        
        # Conteneur pour les éléments
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2196F3'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # Récupérer toutes les classes
        classes = self.db.query(ClassGroup).all()
        
        for class_group in classes:
            # Titre
            elements.append(Paragraph(f"Emploi du temps - {class_group.name}", title_style))
            
            # Créer le tableau
            data = self._create_pdf_table_data(schedule_id, class_group.id)
            
            # Créer le tableau
            table = Table(data, colWidths=[1.5*inch] + [1.8*inch]*6)
            
            # Style du tableau
            table.setStyle(TableStyle([
                # En-tête
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                
                # Corps
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                
                # Colonne des heures
                ('BACKGROUND', (0, 1), (0, -1), colors.lightgrey),
                ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ]))
            
            # Colorier les cellules selon les matières
            self._apply_pdf_colors(table, data, schedule_id, class_group.id)
            
            elements.append(table)
            elements.append(Spacer(1, 0.5*inch))
        
        # Construire le PDF
        doc.build(elements)
        output.seek(0)
        
        return output
    
    def _create_pdf_table_data(self, schedule_id: int, class_id: int) -> List[List[str]]:
        """Créer les données du tableau pour le PDF."""
        # En-tête
        data = [['Heure', 'Dimanche', 'Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi']]
        
        # Périodes
        periods = [
            '08:00-08:45',
            '08:50-09:35',
            '09:40-10:25',
            '10:40-11:25',
            '11:30-12:15',
            '12:20-13:05',
            '13:10-13:55',
            '14:00-14:45'
        ]
        
        # Récupérer les entrées
        entries = self.db.query(ScheduleEntry).filter(
            ScheduleEntry.schedule_id == schedule_id,
            ScheduleEntry.class_id == class_id
        ).all()
        
        # Organiser par jour et période
        schedule_dict = {}
        for entry in entries:
            key = (entry.day, entry.period)
            schedule_dict[key] = entry
        
        # Remplir le tableau
        for period_idx, period_time in enumerate(periods):
            row = [period_time]
            
            for day in ['sunday', 'monday', 'tuesday', 'wednesday', 'thursday', 'friday']:
                entry = schedule_dict.get((day, period_idx))
                
                if entry:
                    subject = self.db.query(Subject).filter(Subject.id == entry.subject_id).first()
                    teacher = self.db.query(Teacher).filter(Teacher.id == entry.teacher_id).first()
                    room = self.db.query(Room).filter(Room.id == entry.room_id).first()
                    
                    cell_text = f"{subject.code}\n{teacher.code}\n{room.code}"
                    row.append(cell_text)
                else:
                    row.append('')
            
            data.append(row)
        
        return data
    
    def _apply_pdf_colors(self, table, data, schedule_id: int, class_id: int):
        """Appliquer les couleurs aux cellules du tableau PDF."""
        # Récupérer les entrées avec leurs matières
        entries = self.db.query(ScheduleEntry).filter(
            ScheduleEntry.schedule_id == schedule_id,
            ScheduleEntry.class_id == class_id
        ).all()
        
        # Créer un dictionnaire des couleurs par position
        color_map = {}
        for entry in entries:
            subject = self.db.query(Subject).filter(Subject.id == entry.subject_id).first()
            if subject and hasattr(subject, 'color') and subject.color:
                day_idx = ['sunday', 'monday', 'tuesday', 'wednesday', 'thursday', 'friday'].index(entry.day) + 1
                row_idx = entry.period + 1
                color_map[(row_idx, day_idx)] = subject.color
        
        # Appliquer les couleurs
        for (row, col), color in color_map.items():
            try:
                table.setStyle(TableStyle([
                    ('BACKGROUND', (col, row), (col, row), colors.HexColor(color)),
                    ('TEXTCOLOR', (col, row), (col, row), colors.white if self._is_dark_color(color) else colors.black)
                ]))
            except:
                pass
    
    def _is_dark_color(self, hex_color: str) -> bool:
        """Déterminer si une couleur est sombre."""
        # Convertir hex en RGB
        hex_color = hex_color.lstrip('#')
        r, g, b = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
        
        # Calculer la luminance
        luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255
        
        return luminance < 0.5
    
    def export_schedule_to_ics(self, schedule_id: int, class_id: Optional[int] = None) -> io.BytesIO:
        """Exporter un emploi du temps au format ICS (iCalendar)."""
        schedule = self.db.query(Schedule).filter(Schedule.id == schedule_id).first()
        if not schedule:
            raise ValueError(f"Schedule {schedule_id} not found")
        
        # Créer un calendrier
        cal = Calendar()
        cal.add('prodid', '-//School Timetable Generator//mxm.dk//')
        cal.add('version', '2.0')
        cal.add('calscale', 'GREGORIAN')
        cal.add('method', 'PUBLISH')
        cal.add('x-wr-calname', f'Emploi du temps - {schedule.name}')
        cal.add('x-wr-timezone', 'Asia/Jerusalem')
        
        # Timezone
        tz = pytz.timezone('Asia/Jerusalem')
        
        # Récupérer les entrées
        query = self.db.query(ScheduleEntry).filter(ScheduleEntry.schedule_id == schedule_id)
        if class_id:
            query = query.filter(ScheduleEntry.class_id == class_id)
        entries = query.all()
        
        # Date de début (prochain dimanche)
        today = datetime.now(tz)
        days_until_sunday = (6 - today.weekday()) % 7
        if days_until_sunday == 0:
            days_until_sunday = 7
        start_date = today.date() + timedelta(days=days_until_sunday)
        
        # Mapper les jours
        day_offset = {
            'sunday': 0,
            'monday': 1,
            'tuesday': 2,
            'wednesday': 3,
            'thursday': 4,
            'friday': 5
        }
        
        # Heures de début et fin des périodes
        period_times = [
            (8, 0, 8, 45),
            (8, 50, 9, 35),
            (9, 40, 10, 25),
            (10, 40, 11, 25),
            (11, 30, 12, 15),
            (12, 20, 13, 5),
            (13, 10, 13, 55),
            (14, 0, 14, 45)
        ]
        
        # Créer les événements
        for entry in entries:
            # Récupérer les informations
            subject = self.db.query(Subject).filter(Subject.id == entry.subject_id).first()
            teacher = self.db.query(Teacher).filter(Teacher.id == entry.teacher_id).first()
            room = self.db.query(Room).filter(Room.id == entry.room_id).first()
            class_group = self.db.query(ClassGroup).filter(ClassGroup.id == entry.class_id).first()
            
            # Calculer la date
            event_date = start_date + timedelta(days=day_offset[entry.day])
            
            # Heures de début et fin
            start_hour, start_min, end_hour, end_min = period_times[entry.period]
            
            # Créer l'événement
            event = Event()
            event.add('summary', f"{subject.name} - {class_group.code}")
            event.add('dtstart', tz.localize(datetime.combine(event_date, datetime.min.time().replace(hour=start_hour, minute=start_min))))
            event.add('dtend', tz.localize(datetime.combine(event_date, datetime.min.time().replace(hour=end_hour, minute=end_min))))
            event.add('location', room.name)
            event.add('description', f"Enseignant: {teacher.first_name} {teacher.last_name}\nMatière: {subject.name}\nClasse: {class_group.name}")
            
            # Récurrence hebdomadaire
            event.add('rrule', {'freq': 'weekly', 'count': 40})  # 40 semaines
            
            # Ajouter au calendrier
            cal.add_component(event)
        
        # Sauvegarder dans un buffer
        output = io.BytesIO()
        output.write(cal.to_ical())
        output.seek(0)
        
        return output 